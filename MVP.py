from csv import writer
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
from send_email import *

web_server = "http://127.0.0.1:5000"
""" Specifying the différent varibable for the devices """
deviceName1 = 'RFCWA0Q3MRV'
deviceName2 = 'RFCT111C17X'
email1 = 'qren2webex@gmail.com'
email2 = 'qren1webex@gmail.com'
token1 = 0
token2 = 0
passed = 0
failed = 0
phoneNumber1 = '0789182615'
phoneNumber2 = '0789182614'
rounds = 0
redFill = PatternFill(start_color= "FF0000", end_color= "FF0000", fill_type="solid")
greenFill = PatternFill(start_color= "00FF00", end_color= "00FF00", fill_type="solid")


class Tests :
    #login
    def MNCQUALIF_10966_in(deviceName,email,token) :
        step1 = requests.post(web_server+'/log-in-bis', json={
            'deviceName': deviceName,
            'email': email,
            'password': '1Sac2billes!',
            'token': token,
            })
        print(step1)
        return ((step1.status_code == 200), step1)

    #logout
    def MNCQUALIF_10966_out(deviceName,token) :
        step1 = requests.post(web_server+'/log-out', json={
            'deviceName': deviceName,
            'token': token,
            })
        return ((step1.status_code == 200), step1)
    
    #call normally
    def MNCQUALIF_11009() :    
        step1 = requests.post(web_server+'/call', json={
            'deviceName': deviceName2,
            'token': token2,
            'destinationNumber': phoneNumber1
            })
        time.sleep(1)
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName1,
            'token': token1,
            'incomingNumber': phoneNumber2
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        return ((step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200), step1, step2, step3)


if __name__ == '__main__' :

    currentTime = datetime.now().time()
    filename = "CampagneBeta-"+str(currentTime)+".xlsx"
    wb = Workbook()
    sheet = wb.active

    response = requests.post(web_server+'/connect', json={'deviceName': deviceName1})
    token1 = response.json()['token']
    response = requests.post(web_server+'/connect', json={'deviceName': deviceName2})
    token2 = response.json()['token']

    """ MNCQUALIF-10996 login/out"""
    MNCQUALIF_10966_in = Tests.MNCQUALIF_10966_in(deviceName2, email2, token2)
    if MNCQUALIF_10966_in[0] :
        sheet.append(("MNCQUALIF-10996 login", "OK"))
        passed+=1
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-10996 login", "KO", MNCQUALIF_10966_in[1].text))
        failed+=1
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill

    time.sleep(5)
    Tests.MNCQUALIF_10966_in(deviceName1, email1, token1)
    time.sleep(5)
    
    while rounds < 10 :
        """ MNCQUALIF-11009 call normally """
        MNCQUALIF_11009 = Tests.MNCQUALIF_11009()
        if MNCQUALIF_11009[0] :
            sheet.append(("MNCQUALIF-11009 call normally", "OK"))
            passed+=1
            for cell in sheet[sheet.max_row] :
                cell.fill = greenFill
        else :
            reason = ""
            for i in range(1,len(MNCQUALIF_11009)) :
                if MNCQUALIF_11009[i].status_code != 200 :
                    reason +=  MNCQUALIF_11009[i].text + " "
            sheet.append(("MNCQUALIF-11009 call normally", "KO", reason))
            failed+=1
            for cell in sheet[sheet.max_row] :
                cell.fill = redFill
        time.sleep(3)

        rounds+=1
    
    """ MNCQUALIF-10996 log out"""
    MNCQUALIF_10966_out = Tests.MNCQUALIF_10966_out(deviceName2, token2) 
    if MNCQUALIF_10966_out[0]:
        sheet.append(("MNCQUALIF-10996 logout", "OK"))
        passed+=1
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-10996 logout", "KO", MNCQUALIF_10966_out[1].text))
        failed+=1
        for cell in sheet[sheet.max_row] :

            cell.fill = redFill

    Tests.MNCQUALIF_10966_out(deviceName1, token1)    

    wb.save(filename=filename)

    service = get_gmail_service()

    if service:
        sender_email = "qlan001webexbeta@gmail.com"
        recipient_email = ["qrenvims1@easymail.orange.com"]
        subject = "Test mail MVP Teaming mobile"
        body = "Ceci est un mail automatique avec en pièce jointe le resultat de la dernière execution ("+ str(passed)+" passed / "+str(failed)+" failed)"
        file_path = filename

        # Créer un message avec pièce jointe
        message = create_message_with_attachment(sender_email, recipient_email, subject, body, file_path)

        # Envoyer le message
        send_message(service, "me", message)