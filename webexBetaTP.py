from csv import writer
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
from send_mail import *

web_server = "http://127.0.0.1:5000"
""" Specifying the différent varibable for the devices """
#deviceName1 = 'R3CR405S9DH'
deviceName1 = 'RFCT111C17X'
deviceName2 = 'RFCWA0Q3MRV'
deviceName3 = 'gy95dqpnkfnr99if'
email1 = 'qlan3webex@gmail.com'
email2 = 'qren1webex@gmail.com'
email3 = 'qlan9webex@gmail.com'
token1 = 0
token2 = 0
token3 = 0
phoneNumberWebexBeta1 = '0789182612'
phoneNumberWebexBeta2 = '0789182614'
phoneNumber3 = '0789182689'
redFill = PatternFill(start_color= "FF0000", end_color= "FF0000", fill_type="solid")
greenFill = PatternFill(start_color= "00FF00", end_color= "00FF00", fill_type="solid")


class Tests :
    #login
    def MNCQUALIF_10966_in(deviceName,email,token) :
        step1 = requests.post(web_server+'/log-in-bis', json={
            'deviceName': deviceName,
            'email': email,
            'password': '1Sac2billes!',
            'token': token,
            })
        return step1.status_code == 200

    #logout
    def MNCQUALIF_10966_out(deviceName,token) :
        step1 = requests.post(web_server+'/log-out', json={
            'deviceName': deviceName,
            'token': token,
            })
        return step1.status_code == 200 

    #check new IM
    def MNCQUALIF_10998() :
        step1 = requests.post(web_server+'/send-im', json={
        "deviceName" : deviceName2,
        "token" : token2,
        "targetMail" : "qlan3webex@gmail.com",
        "instantMessage" : "Ouais"
        })
        time.sleep(10)
        step2 = requests.post(web_server+'/check-new-im', json={
        'deviceName': deviceName1,
        'token': token1,
        'convName': "qRen001 webex"
        })
        return (step1.status_code == 200 and step2.status_code == 200)

    #check new GIM
    def MNCQUALIF_10999() :
        step1 = requests.post(web_server+'/send-gim', json={
            "deviceName" : deviceName2,
            "token" : token2,
            "groupName" : "Test 10999",
            "targetMail" : "qlan3webex@gmail.com",
            "instantMessage" : "Ouais"
            })
        time.sleep(10)
        step2 = requests.post(web_server+'/check-new-gim', json={
            'deviceName': deviceName1,
            'token': token1,
            'convName': "Test 10999"
            })
        return (step1.status_code == 200 and step2.status_code == 200)

    #delete IM  
    def MNCQUALIF_11000() :
        step1 = requests.post(web_server+'/delete-im', json={
            'deviceName': deviceName1,
            'token': token1,
            'convName': "qRen001 webex"
            })
        return(step1.status_code == 200)
    
    #delete GIM
    def MNCQUALIF_11001() :
        step1 = requests.post(web_server+'/delete-gim', json={
            'deviceName': deviceName1,
            'token': token1,
            'convName': "Test 10999"
            })
        return(step1.status_code == 200)
    
    #delete call
    def MNCQUALIF_11004_a() :
        step1 = requests.post(web_server+'/delete-call', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        return(step1.status_code == 200)
    
    #delete all call
    def MNCQUALIF_11004_b() :
        step1 = requests.post(web_server+'/delete-all-call', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        return(step1.status_code == 200)

    #call from logs
    def MNCQUALIF_11005() :
        step1 = requests.post(web_server+'/call-from-logs', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        time.sleep(5)
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName1,
            'token': token1,
            'incomingNumber': phoneNumberWebexBeta2
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        return(step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200)

    #call normally
    def MNCQUALIF_11009() :    
        step1 = requests.post(web_server+'/call', json={
            'deviceName': deviceName2,
            'token': token2,
            'destinationNumber': phoneNumberWebexBeta1
            })
        time.sleep(1)
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName1,
            'token': token1,
            'incomingNumber': phoneNumberWebexBeta2
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        return (step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200)

    #call on hold
    def MNCQUALIF_11011() :
        step1 = requests.post(web_server+'/call', json={
            'deviceName': deviceName2,
            'token': token2,
            'destinationNumber': phoneNumberWebexBeta1
            })
        time.sleep(1)
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName1,
            'token': token1,
            'incomingNumber': phoneNumberWebexBeta2
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/hold', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        time.sleep(5)
        step4 = requests.post(web_server+'/resume', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        time.sleep(5)
        step5 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        return (step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200 and step4.status_code == 200 and step5.status_code == 200)

    #blind transfer
    def MNCQUALIF_11013() :
        step1 = requests.post(web_server+'/call', json={
            'deviceName': deviceName1,
            'token': token1,
            'destinationNumber': phoneNumberWebexBeta2
            })
        time.sleep(1)
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName2,
            'token': token2,
            'incomingNumber': phoneNumberWebexBeta1
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/blind-transfert', json={
            'deviceName': deviceName2,
            'token': token2,
            'transfertTarget' : phoneNumber3
            })
        step4 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName3,
            'token': token3,
            'incomingNumber': phoneNumberWebexBeta1
            }) 
        step5 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        return(step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200 and step4.status_code == 200 and step5.status_code == 200)

    #supervised transfer
    def MNCQUALIF_11014() :
        step1 = requests.post(web_server+'/call', json={
            'deviceName': deviceName1,
            'token': token1,
            'destinationNumber': phoneNumberWebexBeta2
            })
        step2 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName2,
            'token': token2,
            'incomingNumber': phoneNumberWebexBeta1
            })
        time.sleep(5)
        step3 = requests.post(web_server+'/initiate-supervised-transfert', json={
            'deviceName': deviceName2,
            'token': token2,
            'transfertTarget' : phoneNumber3
            })
        time.sleep(5)
        step4 = requests.post(web_server+'/answer', json={
            'deviceName': deviceName3,
            'token': token3,
            'incomingNumber': phoneNumberWebexBeta1
            })
        step5 = requests.post(web_server+'/finalise-supervised-transfert', json={
            'deviceName': deviceName1,
            'token': token1,
            })
        time.sleep(5)
        step6 = requests.post(web_server+'/hang-up', json={
            'deviceName': deviceName2,
            'token': token2,
            })
        return(step1.status_code == 200 and step2.status_code == 200 and step3.status_code == 200 and step4.status_code == 200 and step5.status_code == 200 and step6.status_code == 200) 


if __name__ == '__main__' :

    currentTime = datetime.now().time()
    filename = "CampagneBeta-"+str(currentTime)+".xlsx"
    wb = Workbook()
    sheet = wb.active

    response = requests.post(web_server+'/connect', json={'deviceName': deviceName1})
    token1 = response.json()['token']
    response = requests.post(web_server+'/connect', json={'deviceName': deviceName2})
    token2 = response.json()['token']

    """ MNCQUALIF-10996 login/out"""

    if Tests.MNCQUALIF_10966_in(deviceName2, email2, token2) :
        sheet.append(("MNCQUALIF-10996 login", "OK"))
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-10996 login", "KO"))
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill

    time.sleep(5)
    Tests.MNCQUALIF_10966_in(deviceName1, email1, token1)
    time.sleep(5)
    
    """ MNCQUALIF-10998 check new IM """

    # if Tests.MNCQUALIF_10998() :
    #     sheet.append(("MNCQUALIF-10998 check new IM", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-10998 check new IM", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)
    """ MNCQUALIF-10999 check new GIM """

    # if Tests.MNCQUALIF_10999() :
    #     sheet.append(("MNCQUALIF-10999 check new GIM", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-10999 check new GIM", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)
    """ MNCQUALIF-11000 delete IM """

    # if Tests.MNCQUALIF_11000() :
    #     sheet.append((" MNCQUALIF-11000 delete IM", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append((" MNCQUALIF-11000 delete IM", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)
    """ MNCQUALIF-11001 delete GIM """

    # if Tests.MNCQUALIF_11001() :
    #     sheet.append(("MNCQUALIF-11001 delete GIM", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-11001 delete GIM", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)
    """ MNCQUALIF-11009 call normally """

    if Tests.MNCQUALIF_11009() :
        sheet.append(("MNCQUALIF-11009 call normally", "OK"))
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-11009 call normally", "KO"))
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill
    time.sleep(2)
    """ MNCQUALIF-11011 call on hold"""

    if Tests.MNCQUALIF_11011() :
        sheet.append(("MNCQUALIF-11011 call on hold", "OK"))
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-11011 call on hold", "KO"))
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill
    """ MNCQUALIF-11005 call from logs """

    if Tests.MNCQUALIF_11005() :
        sheet.append(("MNCQUALIF-11005 call from logs", "OK"))
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-11005 call from logs", "KO"))
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill
    time.sleep(2)

    """ MNCQUALIF-11006/12 Meeting video """

    # step1 = requests.post(web_server+'/call', json={
    #     'deviceName': deviceName1,
    #     'token': token1,
    #     'destinationNumber': phoneNumber3
    #     })
    # time.sleep(5)
    # step2 = requests.post(web_server+'/answer', json={
    #         'deviceName': deviceName3,
    #         'token': token3,
    #         'incomingNumber': phoneNumberWebexBeta1
    #         })
    # time.sleep(5)
    # step3 = requests.post(web_server+'/power-up', json={
    #         'deviceName': deviceName1,
    #         'token': token1
    #         })
    # step4 = requests.post(web_server+'/video-call', json={
    #         'deviceName': deviceName1,
    #         'token': token1
    #         })
    # time.sleep(20)
    # step5 = requests.post(web_server+'/end-video-call', json={
    #         'deviceName': deviceName1,
    #         'token': token1
    #         })
    # if step3 :
    #     sheet.append(("MNCQUALIF-11006 power up ", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else : 
    #     sheet.append(("MNCQUALIF-11006 power up ", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # if (step1 and step2 and step3 and step4 and step5) :
    #     sheet.append(("MNCQUALIF-11012 full video call ", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else : 
    #     sheet.append(("MNCQUALIF-11006 full video call ", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    
    
    # """ MNCQUALIF-11013 blind transfer """

    # if Tests.MNCQUALIF_11013() :
    #     sheet.append(("MNCQUALIF-11013", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-11013", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)

    # """ MNCQUALIF-11014 supervised transfer """

    # if Tests.MNCQUALIF_11014() :
    #     sheet.append(("MNCQUALIF-11014", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-11014", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # time.sleep(2)

    """ MNCQUALIF-11004 delete calls"""

    # if Tests.MNCQUALIF_11004_a() :
    #     sheet.append(("MNCQUALIF-11004 one deletion", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    #     sheet.append(("MNCQUALIF-11002 logs update", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    #     sheet.append(("MNCQUALIF-11003 browse log", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-11004 one deletion", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    #     sheet.append(("MNCQUALIF-11002 logs update", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    #     sheet.append(("MNCQUALIF-11003 browse log", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    # if Tests.MNCQUALIF_11004_b() :
    #     sheet.append(("MNCQUALIF-11004 all deletion", "OK"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = greenFill
    # else :
    #     sheet.append(("MNCQUALIF-11004 all deletion", "KO"))
    #     for cell in sheet[sheet.max_row] :
    #         cell.fill = redFill
    time.sleep(2)
    
    """ MNCQUALIF-10996 log out"""

    if Tests.MNCQUALIF_10966_out(deviceName2, token2) :
        sheet.append(("MNCQUALIF-10996 logout", "OK"))
        for cell in sheet[sheet.max_row] :
            cell.fill = greenFill
    else :
        sheet.append(("MNCQUALIF-10996 logout", "KO"))
        for cell in sheet[sheet.max_row] :
            cell.fill = redFill

    Tests.MNCQUALIF_10966_out(deviceName1, token1)    

    wb.save(filename=filename)

sender_email = "qlan001webexbeta@gmail.com"
destinataire_email = "alan.signor.ext@orange.com"
sujet_email = "Sujet de l'email avec pièce jointe"
texte_email = "Voici le corps du message avec une pièce jointe."
chemin_fichier = filename

# Envoi de l'email avec pièce jointe
send_email_with_attachment(sender_email, destinataire_email, sujet_email, texte_email, chemin_fichier)